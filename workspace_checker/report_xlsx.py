"""Styled Excel workbook writer."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import Settings
from .constants import (
    BAND_LIGHT,
    GROUP_TINTS,
    MAX_COL_WIDTH,
    MIN_COL_WIDTH,
    MISSING_FILL,
    MISSING_FONT,
    SEV_COLORS,
    VERDICT_COLORS,
    ZEBRA_FILL,
)
from .models import AuditResult, Severity
from .resolve import build_combined
from .version import __product__, __version__

log = logging.getLogger(__name__)

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NUMERIC_HEADERS = {
    "Num_SymRefs", "Num_ElementTemplates", "Num_ElementTemplateRefs", "Line",
    "Size", "Size_Bytes", "ByLevel_Color", "ByLevel_Weight", "ET_Color", "ET_Weight",
    "Number", "Files", "Count",
}


class WorkbookWriter:
    def __init__(self, settings: Settings):
        self.settings = settings
        self.font = settings.font
        self.brand = settings.get("output", "brand_primary", default="1F3864")
        self.accent = settings.get("output", "brand_accent", default="2E5496")
        self.footer = settings.footer

    # -- primitives --------------------------------------------------------- #
    def _sheet(self, wb: Workbook, name: str) -> Worksheet:
        ws = wb.create_sheet(name[:31])
        ws.sheet_view.showGridLines = False
        return ws

    def _banner(self, ws: Worksheet, title: str, subtitle: str, width: int) -> None:
        width = max(width, 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(name=self.font, size=15, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=self.brand)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 26

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.font = Font(name=self.font, size=10, italic=True, color=self.brand)
        sub.fill = PatternFill("solid", fgColor=BAND_LIGHT)
        sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18

    def _autofit(self, ws: Worksheet, columns: list[str], rows: list[list], start: int = 1) -> None:
        for idx, header in enumerate(columns, start=start):
            longest = len(str(header))
            for row in rows[:2000]:
                if idx - start < len(row):
                    longest = max(longest, len(str(row[idx - start] if row[idx - start] is not None else "")))
            width = min(max(longest + 3, MIN_COL_WIDTH), MAX_COL_WIDTH)
            ws.column_dimensions[get_column_letter(idx)].width = width

    def write_table(
        self,
        ws: Worksheet,
        title: str,
        subtitle: str,
        columns: list[tuple[str, str]],
        rows: list[list],
        status_index: int | None = None,
    ) -> None:
        headers = [c[0] for c in columns]
        self._banner(ws, title, subtitle, len(headers))

        header_row = 3
        for idx, (header, group) in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=idx, value=header)
            cell.font = Font(name=self.font, size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=GROUP_TINTS.get(group, GROUP_TINTS["MISC"]))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER
        ws.row_dimensions[header_row].height = 30

        for r, row in enumerate(rows, start=header_row + 1):
            flagged = (
                status_index is not None
                and status_index < len(row)
                and str(row[status_index] or "").upper() not in ("OK", "")
            )
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.font = Font(
                    name=self.font,
                    size=10,
                    bold=flagged,
                    color=MISSING_FONT if flagged else "000000",
                )
                if flagged:
                    cell.fill = PatternFill("solid", fgColor=MISSING_FILL)
                elif r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=ZEBRA_FILL)
                cell.border = _BORDER
                horizontal = "center" if headers[c - 1] in _NUMERIC_HEADERS else "left"
                cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=False)

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        if rows:
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"
            )
        self._autofit(ws, headers, rows)

    # -- overview ------------------------------------------------------------ #
    def write_overview(self, ws: Worksheet, result: AuditResult) -> None:
        self._banner(
            ws,
            f"{result.tag} - Workspace Standards Health Check",
            f"Generated {result.generated} by {__product__} {__version__}",
            8,
        )
        row = 4

        row = self._verdict_block(ws, row, result)
        row = self._kpi_block(ws, row, result)
        row = self._checks_block(ws, row, "Standards content", result.standards_checks)
        row = self._checks_block(ws, row, "Workspace wiring", result.config_checks)
        row = self._inputs_block(ws, row, result)
        row = self._guide_block(ws, row, result)

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 22
        for col in "CDEFGH":
            ws.column_dimensions[col].width = 20
        ws.cell(row=row + 1, column=1, value=self.footer).font = Font(
            name=self.font, size=9, italic=True, color="808080"
        )

    def _section(self, ws: Worksheet, row: int, text: str) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name=self.font, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=self.accent)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        return row + 1

    def _verdict_block(self, ws: Worksheet, row: int, result: AuditResult) -> int:
        verdict = result.verdict
        fill, color = VERDICT_COLORS.get(verdict, SEV_COLORS["NOT_EVALUATED"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=verdict)
        cell.font = Font(name=self.font, size=14, bold=True, color=color)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 28
        row += 1

        counts = result.issue_counts()
        summary = (
            f"{counts['fail']} failing, {counts['warn']} warning, "
            f"{counts['pass']} passing, {counts['not_evaluated']} not evaluated.  "
            f"Content: {result.standards_verdict}.  Wiring: {result.config_verdict}."
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        note = ws.cell(row=row, column=1, value=summary)
        note.font = Font(name=self.font, size=10, italic=True)
        note.alignment = Alignment(horizontal="center", vertical="center")
        return row + 2

    def _kpi_block(self, ws: Worksheet, row: int, result: AuditResult) -> int:
        row = self._section(ws, row, "Counts")
        counts = result.counts()
        labels = [
            ("Feature definitions", counts["fd"]),
            ("Resolved chain rows", counts["fd_et_level_rows"]),
            ("Element templates", counts["element_templates"]),
            ("Levels", counts["levels"]),
            ("Feature symbologies", counts["feature_symbologies"]),
            ("DGNLIBs", counts["dgnlibs"]),
            ("Config files", counts["cfg_files"]),
            ("Inventoried files", counts["inventory_items"]),
        ]
        for col, (label, value) in enumerate(labels, start=1):
            head = ws.cell(row=row, column=col, value=label)
            head.font = Font(name=self.font, size=9, color="FFFFFF")
            head.fill = PatternFill("solid", fgColor=GROUP_TINTS["MISC"])
            head.alignment = Alignment(horizontal="center", wrap_text=True)
            head.border = _BORDER
            body = ws.cell(row=row + 1, column=col, value=value)
            body.font = Font(name=self.font, size=16, bold=True, color=self.brand)
            body.fill = PatternFill("solid", fgColor=BAND_LIGHT)
            body.alignment = Alignment(horizontal="center")
            body.border = _BORDER
        ws.row_dimensions[row].height = 26
        ws.row_dimensions[row + 1].height = 24
        return row + 3

    def _checks_block(self, ws: Worksheet, row: int, label: str, checks: list) -> int:
        row = self._section(ws, row, f"{label} - findings")
        if not checks:
            ws.cell(row=row, column=1, value="No checks evaluated.").font = Font(
                name=self.font, size=10, italic=True, color="808080"
            )
            return row + 2

        headers = ["Check", "Result", "Outcome", "Detail", "Evidence"]
        spans = [(1, 1), (2, 2), (3, 3), (4, 5), (6, 8)]
        for header, (start, end) in zip(headers, spans):
            if start != end:
                ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
            cell = ws.cell(row=row, column=start, value=header)
            cell.font = Font(name=self.font, size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=GROUP_TINTS["MISC"])
            cell.alignment = Alignment(horizontal="center")
            cell.border = _BORDER
        row += 1

        order = {Severity.FAIL: 0, Severity.WARN: 1, Severity.INFO: 2,
                 Severity.PASS: 3, Severity.NOT_EVALUATED: 4}
        for check in sorted(checks, key=lambda c: order.get(c.severity, 9)):
            fill, color = SEV_COLORS.get(check.severity.value, SEV_COLORS["NOT_EVALUATED"])
            values = [
                (1, 1, check.title),
                (2, 2, check.severity.value),
                (3, 3, check.result),
                (4, 5, check.detail),
                (6, 8, "; ".join(check.evidence[:3])),
            ]
            for start, end, value in values:
                if start != end:
                    ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
                cell = ws.cell(row=row, column=start, value=value)
                cell.font = Font(
                    name=self.font, size=10, bold=(start == 2), color=color if start == 2 else "000000"
                )
                if start == 2:
                    cell.fill = PatternFill("solid", fgColor=fill)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = _BORDER
            row += 1
        return row + 1

    def _inputs_block(self, ws: Worksheet, row: int, result: AuditResult) -> int:
        row = self._section(ws, row, "Inputs detected")
        std = result.standard
        lines = []
        if result.root:
            lines.append(("Workspace root", result.root))
        if std:
            for role in ("FD", "FS", "ET", "LEVEL"):
                present = role in std.inputs_present
                source = std.input_files.get(role, "")
                lines.append(
                    (f"{role} export", source if present else "not provided - dependent checks skipped")
                )
        for entry in (result.config.entry_points if result.config else [])[:5]:
            lines.append(("Config entry point", entry))
        if result.config is not None:
            lines.append(
                (
                    "Configuration scope",
                    "PARTIAL - no workspace root defined, so wiring checks are advisory. "
                    "Supply the parent .cfg or seed roots with --env."
                    if result.config.is_partial
                    else f"Complete - roots defined: {', '.join(result.config.roots_defined)}",
                )
            )
        for message in result.extraction_log[:6]:
            lines.append(("Extraction", message))

        for label, value in lines:
            key = ws.cell(row=row, column=1, value=label)
            key.font = Font(name=self.font, size=10, bold=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            val = ws.cell(row=row, column=2, value=value)
            val.font = Font(name=self.font, size=10)
            val.alignment = Alignment(horizontal="left", wrap_text=False)
            row += 1
        return row + 1

    def _guide_block(self, ws: Worksheet, row: int, result: AuditResult) -> int:
        row = self._section(ws, row, "Sheet guide")
        guide = [
            ("Action_Plan", "Plain-language summary: what is good, what is broken, what to check."),
            ("Config_Health", "Workspace wiring findings with guidance."),
            ("Config_Findings", "One row per individual wiring problem, filterable."),
            ("FD_ET_Level", "One row per feature definition aspect resolved to a template and level."),
            ("FD_Combined", "One summary row per feature definition."),
            ("ElementTemplate_Lookup", "Every element template with its symbology and level."),
            ("Level_Lookup", "The level library with ByLevel symbology and plot flags."),
            ("FS_Lookup", "Feature symbologies and the templates they point at."),
            ("Integrity_Checks", "Every unresolved link. Empty means clean."),
            ("Config_Variables", "Every variable: winner, level, and full definition history."),
            ("Config_Paths", "Each resolved path member and whether it exists."),
            ("Config_DGNLIBs", "Every library, its roles, and any shadowing."),
            ("Inventory", "Everything found in the workspace tree."),
        ]
        for name, description in guide:
            key = ws.cell(row=row, column=1, value=name)
            key.font = Font(name=self.font, size=10, bold=True, color=self.accent)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            val = ws.cell(row=row, column=2, value=description)
            val.font = Font(name=self.font, size=10)
            row += 1
        return row + 1

    # -- action plan ---------------------------------------------------------- #
    def write_action_plan(self, ws: Worksheet, result: AuditResult) -> None:
        self._banner(
            ws,
            f"{result.tag} - What this means and what to do",
            f"{result.verdict}   |   content: {result.standards_verdict}   |   "
            f"wiring: {result.config_verdict}",
            8,
        )
        row = 4
        by_sev = {s: [c for c in result.all_checks if c.severity is s] for s in Severity}

        row = self._plan_intro(ws, row, result)
        row = self._plan_group(
            ws, row, "Must fix before this workspace is used", by_sev[Severity.FAIL],
            "Nothing is failing.",
        )
        row = self._plan_group(ws, row, "Should review", by_sev[Severity.WARN], "No warnings.")
        row = self._plan_group(
            ws, row, "Could not be judged, and why", by_sev[Severity.NOT_EVALUATED],
            "Everything was evaluated.", show_guidance=False,
        )
        self._plan_good(ws, row, by_sev[Severity.PASS])

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 34
        for col in "CDEFGH":
            ws.column_dimensions[col].width = 18

    def _plan_paragraph(self, ws: Worksheet, row: int, text: str, *, italic=False,
                        color="000000", bold=False) -> int:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=2, value=text)
        cell.font = Font(name=self.font, size=10, italic=italic, bold=bold, color=color)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = max(15, 13 * (len(text) // 150 + 1))
        return row + 1

    def _plan_intro(self, ws: Worksheet, row: int, result: AuditResult) -> int:
        counts = result.issue_counts()
        row = self._plan_paragraph(
            ws, row,
            f"{counts['fail']} failing, {counts['warn']} to review, {counts['pass']} passing, "
            f"{counts['not_evaluated']} not evaluated.",
            bold=True,
        )
        if result.config is not None:
            scope = (
                "Only part of the configuration chain was supplied, so wiring findings are "
                "advisory. Re-run with --cfg pointing at the Organization or WorkSpace .cfg, "
                "or seed roots with --env, for a definitive answer."
                if result.config.is_partial
                else "The configuration chain looks complete (roots: "
                     f"{', '.join(result.config.roots_defined)})."
            )
            row = self._plan_paragraph(ws, row, scope, italic=True, color="595959")
        return row + 1

    def _plan_group(self, ws: Worksheet, row: int, heading: str, checks: list,
                    empty_text: str, show_guidance: bool = True) -> int:
        row = self._section(ws, row, f"{heading}  ({len(checks)})")
        if not checks:
            return self._plan_paragraph(ws, row, empty_text, italic=True, color="595959") + 1

        for check in checks:
            head = ws.cell(row=row, column=2, value=check.title)
            head.font = Font(name=self.font, size=11, bold=True, color=self.brand)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
            res = ws.cell(row=row, column=3, value=check.result)
            res.font = Font(name=self.font, size=10, bold=True)
            res.alignment = Alignment(horizontal="left", wrap_text=True)
            row += 1

            if check.detail:
                row = self._plan_paragraph(ws, row, check.detail, color="404040")
            if show_guidance and check.guidance:
                row = self._plan_paragraph(
                    ws, row, f"What to do:  {check.guidance}", italic=True, color="1F6F8B"
                )
            for finding in check.findings[:8]:
                row = self._plan_paragraph(ws, row, f"    - {finding.summary}", color="595959")
            extra = len(check.findings) - 8
            if extra > 0:
                row = self._plan_paragraph(
                    ws, row, f"    ... and {extra} more, see Config_Findings",
                    italic=True, color="808080",
                )
            row += 1
        return row

    def _plan_good(self, ws: Worksheet, row: int, checks: list) -> int:
        row = self._section(ws, row, f"Confirmed good  ({len(checks)})")
        if not checks:
            return self._plan_paragraph(ws, row, "Nothing passed.", italic=True, color="595959")
        for check in checks:
            key = ws.cell(row=row, column=2, value=check.title)
            key.font = Font(name=self.font, size=10, bold=True, color="375623")
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
            val = ws.cell(row=row, column=3, value=check.result)
            val.font = Font(name=self.font, size=10)
            row += 1
        return row + 1


# --------------------------------------------------------------------------- #
# Sheet builders
# --------------------------------------------------------------------------- #
def _fd_et_level(result: AuditResult):
    columns = [
        ("FD_Name", "FD"), ("FD_ItemType", "FD"), ("FD_Path", "FD"),
        ("SymbologyType", "FS"), ("FS_FeaturePath", "FS"), ("FS_Name", "FS"),
        ("ET_Relationship", "ET"), ("ElementTemplate_Path", "ET"), ("ElementTemplate_Name", "ET"),
        ("Level", "LV"), ("ByLevel_Color", "LV"), ("ByLevel_Weight", "LV"), ("ByLevel_Style", "LV"),
        ("ET_Color", "ET"), ("ET_Weight", "ET"), ("ET_LineStyle", "ET"),
        ("Material", "ET"), ("TextStyle", "ET"), ("Status", "MISC"),
    ]
    rows = [
        [r.fd_name, r.fd_item_type, r.fd_path, r.stype, r.fs_featurepath, r.fs_name,
         r.et_relationship, r.et_path, r.et_name, r.level, r.bylevel_color, r.bylevel_weight,
         r.bylevel_style, r.et_color, r.et_weight, r.et_linestyle, r.material, r.textstyle, r.status]
        for r in result.rows
    ]
    return columns, rows, len(columns) - 1


def _fd_combined(result: AuditResult):
    columns = [
        ("FD_Name", "FD"), ("FD_ItemType", "FD"), ("FD_Path", "FD"),
        ("SymbologyTypes", "FS"), ("Num_SymRefs", "FS"), ("Num_ElementTemplates", "ET"),
        ("Distinct_Levels", "LV"), ("Distinct_ByLevel_Colors", "LV"), ("Aspect_Level_Summary", "MISC"),
    ]
    records = build_combined(result.standard, result.rows) if result.standard else []
    rows = [[rec[c[0]] for c in columns] for rec in records]
    return columns, rows, None


def _et_lookup(result: AuditResult):
    columns = [
        ("ElementTemplate_Path", "ET"), ("ElementTemplate_Name", "ET"), ("Level", "LV"),
        ("ByLevel_Color", "LV"), ("ByLevel_Weight", "LV"), ("ByLevel_Style", "LV"), ("Plot", "LV"),
        ("ET_Color", "ET"), ("ET_Weight", "ET"), ("ET_LineStyle", "ET"), ("Material", "ET"),
        ("TextStyle", "ET"), ("ElementClass", "ET"), ("Transparency", "ET"),
    ]
    std = result.standard
    rows = []
    for tpl in (std.et.values() if std else []):
        level = std.levels.get(tpl.level)
        rows.append([
            tpl.et_path, tpl.name, tpl.level,
            level.bylevel_color if level else "", level.bylevel_weight if level else "",
            level.bylevel_style if level else "", level.plot if level else "",
            tpl.color, tpl.weight, tpl.linestyle, tpl.material, tpl.textstyle,
            tpl.element_class, tpl.transparency,
        ])
    return columns, rows, None


def _level_lookup(result: AuditResult):
    columns = [
        ("Level_Name", "LV"), ("Number", "LV"), ("Description", "LV"), ("ByLevel_Color", "LV"),
        ("ByLevel_Weight", "LV"), ("ByLevel_Style", "LV"), ("Plot", "LV"),
        ("GlobalDisplay", "LV"), ("GlobalFreeze", "LV"),
    ]
    std = result.standard
    rows = [
        [lv.name, lv.number, lv.description, lv.bylevel_color, lv.bylevel_weight,
         lv.bylevel_style, lv.plot, lv.global_display, lv.global_freeze]
        for lv in (std.levels.values() if std else [])
    ]
    return columns, rows, None


def _fs_lookup(result: AuditResult):
    columns = [
        ("SymbologyType", "FS"), ("FS_FeaturePath", "FS"), ("FS_Name", "FS"),
        ("Num_ElementTemplateRefs", "ET"), ("ElementTemplateRefs", "ET"),
    ]
    std = result.standard
    rows = [
        [fs.stype, fs.featurepath, fs.name, len(fs.refs),
         "; ".join(f"{r.relationship}={r.et_path}" for r in fs.refs)]
        for fs in (std.fs.values() if std else [])
    ]
    return columns, rows, None


def _integrity(result: AuditResult):
    columns = [
        ("FD_Name", "FD"), ("SymbologyType", "FS"), ("FS_FeaturePath", "FS"),
        ("FS_Name", "FS"), ("Issue", "MISC"), ("Detail", "MISC"),
    ]
    rows = [
        [i.fd_name, i.stype, i.fs_featurepath, i.fs_name, i.issue, i.detail]
        for i in result.issues
    ]
    return columns, rows, None


def _config_variables(result: AuditResult):
    columns = [
        ("Name", "CFG"), ("Effective_Value", "CFG"), ("Level_Won", "CFG"), ("Locked", "CFG"),
        ("Members", "CFG"), ("Members_Exist", "CFG"), ("Definition_History", "CFG"),
    ]
    model = result.config
    if model is None:
        return columns, [], None
    exists_by_var: dict[str, list[bool]] = {}
    for member in model.path_members:
        exists_by_var.setdefault(member.variable, []).append(member.exists)

    rows = []
    for var in sorted(model.variables.values(), key=lambda v: v.name):
        flags = exists_by_var.get(var.name, [])
        rows.append([
            var.name, var.value, var.level, "yes" if var.locked else "",
            len(var.members),
            "".join("Y" if f else "N" for f in flags),
            " | ".join(
                f"{d.level} {d.origin} {d.operator} {d.raw_value.strip()}"
                + ("" if d.applied else f"  [{d.note}]")
                for d in var.history
            ),
        ])
    return columns, rows, None


def _config_paths(result: AuditResult):
    columns = [
        ("Variable", "CFG"), ("Member_Path", "CFG"), ("Resolved", "CFG"),
        ("Exists", "CFG"), ("Undefined_Variables", "CFG"),
        ("Source_File", "CFG"), ("Line", "CFG"), ("Status", "MISC"),
    ]
    model = result.config
    rows = [
        [m.variable, m.member, m.resolved, "yes" if m.exists else "no",
         ", ".join(m.unresolved),
         Path(m.source_file).name if m.source_file else "", m.line, m.status]
        for m in (model.path_members if model else [])
    ]
    return columns, rows, len(columns) - 1


def _config_dgnlibs(result: AuditResult):
    columns = [
        ("Path", "CFG"), ("Exists", "CFG"), ("Roles_Provided", "CFG"),
        ("Precedence_Level", "CFG"), ("On_Config", "CFG"), ("Shadowed_By", "CFG"),
        ("Note", "MISC"), ("Status", "MISC"),
    ]
    rows = []
    for lib in result.dgnlibs:
        status = "OK"
        if not lib.exists:
            status = "MISSING"
        elif lib.shadowed_by:
            status = "SHADOWED"
        elif not lib.on_config:
            status = "UNWIRED"
        rows.append([
            lib.path, "yes" if lib.exists else "no", ", ".join(lib.roles),
            lib.precedence_level, "yes" if lib.on_config else "no",
            lib.shadowed_by, lib.note, status,
        ])
    return columns, rows, len(columns) - 1


def _config_findings(result: AuditResult):
    columns = [
        ("Check", "CFG"), ("Outcome", "CFG"), ("Item", "CFG"), ("Detail", "CFG"),
        ("Source_File", "CFG"), ("Line", "CFG"), ("What_To_Do", "MISC"), ("Status", "MISC"),
    ]
    rows = []
    for check in result.config_checks:
        for finding in check.findings:
            rows.append([
                check.title, check.severity.value, finding.item, finding.detail,
                Path(finding.source_file).name if finding.source_file else "",
                finding.line or "", check.guidance,
                "OK" if check.severity is Severity.PASS else check.severity.value,
            ])
    return columns, rows, len(columns) - 1


def _inventory(result: AuditResult):
    columns = [
        ("Relative_Path", "INV"), ("Kind", "INV"), ("Size_Bytes", "INV"),
        ("Modified", "INV"), ("SHA1", "INV"), ("Note", "MISC"),
    ]
    tree = result.tree
    rows = [
        [i.relpath, i.kind, i.size, i.modified, i.sha1, i.note]
        for i in (tree.inventory if tree else [])
    ]
    return columns, rows, None


_SHEETS = [
    ("Config_Findings", "Configuration findings", "One row per wiring problem, with what to do.", _config_findings),
    ("FD_ET_Level", "Resolved symbology chain", "FD aspect to element template to level.", _fd_et_level),
    ("FD_Combined", "Feature definition summary", "One row per feature definition.", _fd_combined),
    ("ElementTemplate_Lookup", "Element templates", "Every template with its symbology.", _et_lookup),
    ("Level_Lookup", "Level library", "ByLevel symbology and plot flags.", _level_lookup),
    ("FS_Lookup", "Feature symbologies", "Templates referenced by each symbology.", _fs_lookup),
    ("Integrity_Checks", "Integrity issues", "Empty means every reference resolves.", _integrity),
    ("Config_Variables", "Configuration variables", "Effective value with full provenance.", _config_variables),
    ("Config_Paths", "Configuration paths", "Every resolved path member and its status.", _config_paths),
    ("Config_DGNLIBs", "DGNLIB inventory", "Roles, precedence and shadowing.", _config_dgnlibs),
    ("Inventory", "Workspace inventory", "Everything found under the workspace root.", _inventory),
]


def write_workbook(result: AuditResult, out_path: str | Path, settings: Settings) -> Path:
    """Write the full review workbook and return its path."""
    writer = WorkbookWriter(settings)
    wb = Workbook()
    wb.remove(wb.active)

    overview = writer._sheet(wb, "Overview")
    writer.write_overview(overview, result)

    plan = writer._sheet(wb, "Action_Plan")
    writer.write_action_plan(plan, result)

    health = writer._sheet(wb, "Config_Health")
    columns = [("Check", "CFG"), ("Outcome", "CFG"), ("Result", "CFG"),
               ("Detail", "MISC"), ("What_To_Do", "MISC"), ("Evidence", "MISC")]
    rows = [
        [c.title, c.severity.value, c.result, c.detail, c.guidance, "; ".join(c.evidence[:5])]
        for c in result.config_checks
    ]
    writer.write_table(
        health,
        f"{result.tag} - Workspace wiring health",
        f"Configuration verdict: {result.config_verdict}",
        columns,
        rows,
    )

    for name, title, subtitle, builder in _SHEETS:
        columns, rows, status_index = builder(result)
        ws = writer._sheet(wb, name)
        writer.write_table(ws, f"{result.tag} - {title}", subtitle, columns, rows, status_index)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log.info("Wrote %s", out)
    return out


def default_workbook_name(tag: str) -> str:
    return f"{tag}_Workspace_Health_Review.xlsx"


def timestamp() -> str:
    return datetime.now().isoformat(timespec="seconds")
